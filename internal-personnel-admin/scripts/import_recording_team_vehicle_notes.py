"""Copy vehicle plates from selected recording-team tabs into internal notes."""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", type=Path, required=True)
    parser.add_argument("--team-sheet", type=Path, action="append", required=True)
    args = parser.parse_args()
    connection = sqlite3.connect(args.database)
    updated = 0
    for workbook_path in args.team_sheet:
        sheet = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True).active
        headings = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        name_column, plate_column = headings.index("姓名"), headings.index("車號")
        for row in sheet.iter_rows(min_row=2):
            name = str(row[name_column].value or "").strip()
            plate = str(row[plate_column].value or "").strip()
            if not name or not plate:
                continue
            person = connection.execute("SELECT id, notes FROM people WHERE display_name = ?", (name,)).fetchone()
            if not person:
                raise RuntimeError(f"Expected person not found: {name}")
            note = f"車號：{plate}"
            prior = str(person[1] or "").strip()
            if note in prior:
                continue
            notes = f"{prior}\n{note}" if prior else note
            connection.execute("UPDATE people SET notes = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (notes, person[0]))
            connection.execute("INSERT INTO audit_log (person_id, action, actor_email) VALUES (?, 'imported-vehicle-plate', 'directory-correction')", (person[0],))
            updated += 1
    connection.commit()
    connection.close()
    print(f"vehicle_notes_updated={updated}")


if __name__ == "__main__":
    main()
