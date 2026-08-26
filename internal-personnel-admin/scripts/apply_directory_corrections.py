"""Apply confirmed name merges and assistant relationships to the local directory."""

from __future__ import annotations

import argparse
import sqlite3
import uuid
from pathlib import Path

import openpyxl


def person(connection: sqlite3.Connection, name: str) -> sqlite3.Row:
    row = connection.execute("SELECT * FROM people WHERE display_name = ?", (name,)).fetchone()
    if not row:
        raise RuntimeError(f"Expected person not found: {name}")
    return row


def set_role(connection: sqlite3.Connection, person_id: str, title: str) -> None:
    connection.execute("DELETE FROM person_roles WHERE person_id = ?", (person_id,))
    connection.execute("INSERT INTO person_roles (person_id, role_name) VALUES (?, ?)", (person_id, title))


def import_recording_assistants(connection: sqlite3.Connection, workbook_path: Path) -> int:
    sheet = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True).active
    header = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_column, id_column, phone_column = (header.index(value) for value in ("姓名", "身分證字號", "電話"))
    count = 0
    for row in sheet.iter_rows(min_row=2):
        name = str(row[name_column].value or "").strip()
        if not name or name == "陳凱馨":
            continue
        identity_number = str(row[id_column].value or "").strip()
        phone = str(row[phone_column].value or "").strip()
        if phone[:1].lower() == "o":
            phone = f"0{phone[1:]}"
        found = connection.execute("SELECT id FROM people WHERE display_name = ?", (name,)).fetchone()
        if found:
            person_id = str(found[0])
            connection.execute("UPDATE people SET phone = COALESCE(NULLIF(phone, ''), ?), id_document_type = COALESCE(NULLIF(id_document_type, ''), '身分證'), id_document_number = COALESCE(NULLIF(id_document_number, ''), ?) WHERE id = ?", (phone, identity_number, person_id))
        else:
            person_id = str(uuid.uuid4())
            connection.execute("INSERT INTO people (id, display_name, legal_name_zh, phone, id_document_type, id_document_number, id_document_status, is_active) VALUES (?, ?, ?, ?, '身分證', ?, 'missing', 1)", (person_id, name, name, phone, identity_number))
        set_role(connection, person_id, "陳凱馨錄音助理")
        connection.execute("INSERT INTO audit_log (person_id, action, actor_email) VALUES (?, 'imported-recording-assistant', 'directory-correction')", (person_id,))
        count += 1
    return count


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", type=Path, required=True)
    parser.add_argument("--recording-team", type=Path, action="append", required=True)
    args = parser.parse_args()
    connection = sqlite3.connect(args.database)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    sophie, wang = person(connection, "Sophie Wang"), person(connection, "王子欣")
    connection.execute("UPDATE people SET legal_name_zh = '王子欣', legal_name_en = 'Sophie Wang' WHERE id = ?", (wang["id"],))
    for role in connection.execute("SELECT role_name FROM person_roles WHERE person_id = ?", (sophie["id"],)):
        connection.execute("INSERT OR IGNORE INTO person_roles (person_id, role_name) VALUES (?, ?)", (wang["id"], str(role[0])))
    connection.execute("DELETE FROM people WHERE id = ?", (sophie["id"],))
    set_role(connection, person(connection, "陳婷怡")["id"], "江國安錄音助理")
    set_role(connection, person(connection, "許夢芬")["id"], "黃俊綸攝影助理")
    assistant_count = sum(import_recording_assistants(connection, path) for path in args.recording_team)
    connection.commit()
    total = connection.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    connection.close()
    print(f"merged_sophie_wang=1 recording_assistant_rows={assistant_count} total_people={total}")


if __name__ == "__main__":
    main()
