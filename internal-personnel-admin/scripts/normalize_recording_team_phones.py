"""Correct the Google Sheet's letter-o placeholder to a string telephone zero."""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", type=Path, required=True)
    parser.add_argument("--team-sheet", type=Path, action="append", required=True)
    args = parser.parse_args()
    connection = sqlite3.connect(args.database)
    updated = 0
    for path in args.team_sheet:
        sheet = openpyxl.load_workbook(path, data_only=True, read_only=True).active
        headings = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        name_column, phone_column = headings.index("姓名"), headings.index("電話")
        for row in sheet.iter_rows(min_row=2):
            name, phone = str(row[name_column].value or "").strip(), str(row[phone_column].value or "").strip()
            if not name or phone[:1].lower() != "o":
                continue
            corrected_phone = f"0{phone[1:]}"
            person = connection.execute("SELECT id FROM people WHERE display_name = ?", (name,)).fetchone()
            if not person:
                raise RuntimeError(f"Expected person not found: {name}")
            connection.execute("UPDATE people SET phone = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (corrected_phone, person[0]))
            connection.execute("INSERT INTO audit_log (person_id, action, actor_email) VALUES (?, 'normalized-phone-prefix', 'directory-correction')", (person[0],))
            updated += 1
    connection.commit()
    connection.close()
    print(f"recording_team_phone_prefixes_corrected={updated}")


if __name__ == "__main__":
    main()
