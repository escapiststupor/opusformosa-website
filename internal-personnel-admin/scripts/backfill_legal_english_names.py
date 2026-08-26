"""Fill blank legal English names from the original musician-form export.

Existing nonblank values are deliberately preserved: later staff corrections are
more authoritative than a form import.
"""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import openpyxl


EMAIL = "電子郵件地址"
ENGLISH_NAME = "英文姓名（如適用，與護照一致）｜Name in English (if applicable, as shown on your passport)"


def text(value: object) -> str:
    return str(value or "").strip()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--database", type=Path, required=True)
    args = parser.parse_args()
    sheet = openpyxl.load_workbook(args.source, data_only=True, read_only=True).active
    headings = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    email_column, english_column = headings.index(EMAIL), headings.index(ENGLISH_NAME)
    connection = sqlite3.connect(args.database)
    updated = 0
    for row in sheet.iter_rows(min_row=2):
        email, english_name = text(row[email_column].value).lower(), text(row[english_column].value)
        if not email or not english_name:
            continue
        record = connection.execute("SELECT id, legal_name_en FROM people WHERE lower(email) = ?", (email,)).fetchone()
        if record and not text(record[1]):
            connection.execute("UPDATE people SET legal_name_en = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (english_name, record[0]))
            connection.execute("INSERT INTO audit_log (person_id, action, actor_email) VALUES (?, 'backfilled-legal-english-name', 'directory-correction')", (record[0],))
            updated += 1
    connection.commit()
    connection.close()
    print(f"legal_english_names_backfilled={updated}")


if __name__ == "__main__":
    main()
