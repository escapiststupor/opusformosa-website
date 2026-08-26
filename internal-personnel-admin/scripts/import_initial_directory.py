"""Import the verified musician-form export plus named internal contacts.

This keeps every identifier exactly as exported. It deliberately does not infer
leading zeroes, combine postal codes with account numbers, or mark an ID document
as received merely because its number was entered in the form.
"""

from __future__ import annotations

import argparse
import re
import sqlite3
import uuid
from pathlib import Path

import openpyxl


SOURCE_HEADERS = {
    "name": "姓名（與證件一致）｜Full legal name (as shown on your identification)",
    "legal_name_en": "英文姓名（如適用，與護照一致）｜Name in English (if applicable, as shown on your passport)",
    "role": "職務／演奏樂器｜Role / Instrument",
    "email": "電子郵件地址",
    "phone": "聯絡電話｜Phone number",
    "nationality": "國籍｜Nationality",
    "residency": "國籍及在台居住狀態｜Nationality and Taiwan residency status",
    "id_type": "證件類型｜Type of identification",
    "id_number": "證件號碼｜Identification number",
    "bank_name": "銀行名稱｜Bank name",
    "bank_code": "銀行代碼｜Bank code",
    "branch_name": "分行名稱｜Branch name",
    "holder": "戶名｜Account holder’s name",
    "account": "銀行帳號｜Bank account number",
}

# These people were named in the event brief but did not appear in the musician form.
# Empty contact/identity/payment fields are intentional: they must be collected later.
EXTRA_PEOPLE = [
    ("小林愛実", "", "", "鋼琴"),
    ("黃凱珉", "", "Sirena Huang", "小提琴"),
    ("Adrien La Marca", "", "", "中提琴"),
    ("Jinjoo Cho", "", "", "小提琴"),
    ("Edgar Moreau", "", "", "大提琴"),
    ("Kyu Yeon Kim", "", "", "鋼琴"),
    ("Boris Borgolotto", "", "", "小提琴"),
    ("Brannon Cho", "", "", "大提琴"),
    ("Sophie Wang", "", "", "小提琴"),
    ("江國安", "", "", "錄音師"),
    ("陳婷怡", "", "", "錄音師助理"),
    ("呂少評", "", "", "舞台監督"),
    ("巨彥博", "", "", "場務"),
    ("陳玥絨", "", "", "攝影師"),
    ("許夢芬", "", "", "攝影助理"),
    ("王子欣", "", "", "小提琴"),
    ("黃慈恩", "", "", "場務"),
    ("馬竟家", "", "", "場務"),
]


def as_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized(value: str) -> str:
    return re.sub(r"\s+", "", value).casefold()


def find_person(connection: sqlite3.Connection, name: str, email: str, id_number: str) -> str | None:
    if email:
        row = connection.execute("SELECT id FROM people WHERE lower(email) = lower(?)", (email,)).fetchone()
        if row:
            return str(row[0])
    if id_number:
        row = connection.execute("SELECT id FROM people WHERE id_document_number = ?", (id_number,)).fetchone()
        if row:
            return str(row[0])
    for row in connection.execute("SELECT id, display_name FROM people"):
        if normalized(str(row[1])) == normalized(name):
            return str(row[0])
    return None


def insert_person(connection: sqlite3.Connection, *, name: str, legal_name_zh: str, legal_name_en: str, email: str = "", phone: str = "", nationality: str = "", residency: str = "", id_type: str = "", id_number: str = "", bank_name: str = "", bank_code: str = "", branch_name: str = "", holder: str = "", account: str = "", role: str = "") -> bool:
    person_id = find_person(connection, name, email, id_number)
    if person_id:
        return False
    person_id = str(uuid.uuid4())
    connection.execute(
        """INSERT INTO people (
          id, display_name, legal_name_zh, legal_name_en, email, phone, nationality,
          residency_status, id_document_type, id_document_number, id_document_status,
          bank_name, bank_code, bank_branch, bank_account_holder, bank_account_number, is_active
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'missing', ?, ?, ?, ?, ?, 1)""",
        (person_id, name, legal_name_zh, legal_name_en, email, phone, nationality, residency, id_type, id_number, bank_name, bank_code, branch_name, holder, account),
    )
    if role:
        connection.execute("INSERT INTO person_roles (person_id, role_name) VALUES (?, ?)", (person_id, role))
    connection.execute("INSERT INTO audit_log (person_id, action, actor_email) VALUES (?, 'imported', 'initial-directory-import')", (person_id,))
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--database", type=Path, required=True)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.source, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [as_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    columns = {field: headers.index(label) for field, label in SOURCE_HEADERS.items()}

    connection = sqlite3.connect(args.database)
    connection.execute("PRAGMA foreign_keys = ON")
    imported_form = 0
    for row in sheet.iter_rows(min_row=2):
        values = {field: as_text(row[index].value) for field, index in columns.items()}
        if not values["name"]:
            continue
        # The original legal-name field is retained verbatim, including mixed-script names.
        if insert_person(connection, name=values["name"], legal_name_zh=values["name"], legal_name_en=values["legal_name_en"], email=values["email"], phone=values["phone"], nationality=values["nationality"], residency=values["residency"], id_type=values["id_type"], id_number=values["id_number"], bank_name=values["bank_name"], bank_code=values["bank_code"], branch_name=values["branch_name"], holder=values["holder"], account=values["account"], role=values["role"]):
            imported_form += 1
    imported_extra = 0
    for display_name, legal_name_zh, legal_name_en, role in EXTRA_PEOPLE:
        if insert_person(connection, name=display_name, legal_name_zh=legal_name_zh, legal_name_en=legal_name_en, role=role):
            imported_extra += 1
    connection.commit()
    total = connection.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    connection.close()
    print(f"imported_form={imported_form} imported_named_contacts={imported_extra} total_people={total}")


if __name__ == "__main__":
    main()
