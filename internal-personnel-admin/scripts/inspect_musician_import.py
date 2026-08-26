"""Read-only preflight checks for the Google Form musician export.

All bank identifiers stay as source strings. The report flags ambiguity; it does not
alter source values or import any record.
"""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


SOURCE = Path("/private/tmp/opus-musicians-source.xlsx")
REPORT = Path("/private/tmp/opus-musicians-preflight.json")

FIELDS = {
    "name": "姓名（與證件一致）｜Full legal name (as shown on your identification)",
    "role": "職務／演奏樂器｜Role / Instrument",
    "email": "電子郵件地址",
    "phone": "聯絡電話｜Phone number",
    "nationality": "國籍｜Nationality",
    "residency": "國籍及在台居住狀態｜Nationality and Taiwan residency status",
    "id_type": "證件類型｜Type of identification",
    "id_number": "證件號碼｜Identification number",
    "bank_name": "銀行名稱｜Bank name",
    "bank_code": "銀行代碼｜Bank code",
    "branch_name": "分行名稱｜Branch name",
    "branch_code": "分行代碼｜Branch code",
    "holder": "戶名｜Account holder’s name",
    "account": "銀行帳號｜Bank account number",
    "document_confirmation": "證件寄送確認｜Identification document confirmation",
}


def as_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def mask(value: str) -> str:
    if not value:
        return ""
    return "•" * max(0, len(value) - 4) + value[-4:]


def main() -> None:
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [as_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    columns = {key: headers.index(label) for key, label in FIELDS.items()}

    records: list[dict[str, object]] = []
    numeric_identifier_cells: list[dict[str, object]] = []
    account_to_records: defaultdict[str, list[dict[str, object]]] = defaultdict(list)
    identity_to_records: defaultdict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    document_statuses: Counter[str] = Counter()
    required = ("name", "role", "id_type", "id_number", "bank_name", "account")

    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {key: as_text(row[index].value) for key, index in columns.items()}
        if not any(values.values()):
            continue
        record = {"row": row_number, **values}
        records.append(record)
        document_statuses[values["document_confirmation"] or "（空白）"] += 1
        for field in ("phone", "id_number", "bank_code", "branch_code", "account"):
            cell = row[columns[field]]
            if cell.value not in (None, "") and not isinstance(cell.value, str):
                numeric_identifier_cells.append(
                    {
                        "row": row_number,
                        "name": values["name"],
                        "field": field,
                        "source_type": type(cell.value).__name__,
                        "source_value": mask(values[field]) if field in {"account", "id_number", "phone"} else values[field],
                    }
                )
        if values["account"]:
            account_to_records[values["account"]].append(record)
        if values["id_type"] and values["id_number"]:
            identity_to_records[(values["id_type"], values["id_number"])].append(record)

    missing_required = [
        {"row": item["row"], "name": item["name"], "missing": [field for field in required if not item[field]]}
        for item in records
        if any(not item[field] for field in required)
    ]
    duplicate_accounts = [
        {"account_last4": mask(account), "people": [{"row": item["row"], "name": item["name"]} for item in people]}
        for account, people in account_to_records.items()
        if len(people) > 1
    ]
    duplicate_identities = [
        {"id_type": key[0], "people": [{"row": item["row"], "name": item["name"]} for item in people]}
        for key, people in identity_to_records.items()
        if len(people) > 1
    ]
    post_office_rows = [
        {
            "row": item["row"],
            "name": item["name"],
            "bank_name": item["bank_name"],
            "bank_code": item["bank_code"],
            "branch_name": item["branch_name"],
            "branch_code": item["branch_code"],
            "account_last4": mask(item["account"]),
            "account_length": len(item["account"]),
        }
        for item in records
        if "郵局" in item["bank_name"] or item["bank_code"] == "700"
    ]
    names_with_latin_characters = [
        {"row": item["row"], "name": item["name"]}
        for item in records
        if any(char.isascii() and char.isalpha() for char in item["name"])
    ]
    short_or_nonstandard_branch_codes = [
        {"row": item["row"], "name": item["name"], "branch_code": item["branch_code"]}
        for item in records
        if item["branch_code"] and (len(item["branch_code"]) != 4 or not item["branch_code"].isdigit())
    ]

    report = {
        "records": len(records),
        "source_sheet": sheet.title,
        "document_confirmation_counts": dict(document_statuses),
        "missing_required": missing_required,
        "numeric_identifier_cells": numeric_identifier_cells,
        "duplicate_accounts": duplicate_accounts,
        "duplicate_identities": duplicate_identities,
        "post_office_rows": post_office_rows,
        "names_with_latin_characters": names_with_latin_characters,
        "short_or_nonstandard_branch_codes": short_or_nonstandard_branch_codes,
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
