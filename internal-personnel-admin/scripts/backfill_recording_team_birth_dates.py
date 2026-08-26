"""Fill birth dates and string phone values from selected recording-team tabs."""

from __future__ import annotations

import argparse
import sqlite3
from datetime import date, datetime
from pathlib import Path

import openpyxl


def text(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value or "").strip()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", type=Path, required=True)
    parser.add_argument("--team-sheet", type=Path, action="append", required=True)
    args = parser.parse_args()
    connection = sqlite3.connect(args.database)
    updated_people: set[str] = set()
    for path in args.team_sheet:
        sheet = openpyxl.load_workbook(path, data_only=True, read_only=True).active
        headings = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        name_column, birth_column, phone_column = (headings.index(label) for label in ("姓名", "出生年月日", "電話"))
        for row in sheet.iter_rows(min_row=2):
            name, birth_date, phone = text(row[name_column].value), text(row[birth_column].value), text(row[phone_column].value)
            if not name:
                continue
            if phone[:1].lower() == "o":
                phone = f"0{phone[1:]}"
            person = connection.execute("SELECT id FROM people WHERE display_name = ?", (name,)).fetchone()
            if not person:
                raise RuntimeError(f"Expected person not found: {name}")
            connection.execute("UPDATE people SET birth_date = ?, phone = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (birth_date, phone, person[0]))
            updated_people.add(name)
    connection.commit()
    connection.close()
    print(f"recording_team_birth_dates_backfilled={len(updated_people)}")


if __name__ == "__main__":
    main()
