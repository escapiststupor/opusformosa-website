from pathlib import Path

from openpyxl import load_workbook


workbook_path = Path('outputs/opus_schedule_update/Opus_Formosa_2026_Rehearsal_Schedule_UPDATED.xlsx')
workbook = load_workbook(workbook_path, data_only=False, read_only=True)
print(workbook.sheetnames)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    matches = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if any('Brannon' in str(value) for value in row if value is not None):
            matches.append((row_number, row))
    if matches:
        print(f'--- {sheet_name}')
        for row_number, row in matches:
            print(row_number, list(row))

sheet = workbook['Individual Schedules']
print('--- Individual Schedules: Brannon detail')
for row_number in range(142, 158):
    print(row_number, [sheet.cell(row_number, column).value for column in range(1, 8)])
