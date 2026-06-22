#!/usr/bin/env python3
"""
Generate NCH seat-map CSV from blank Excel.

Extracts the full cell-detection logic from color_nch.py and instead of
applying colors, outputs a CSV mapping Excel cell positions to seat identities.

Output columns: excel_row,col,section,row,seat
"""

import csv
import sys
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
EXCEL_SRC = Path('/Users/pyen/OpusFormosa/ticketing/2026/0904NCH/國家音樂廳座位圖1150303(多元友善席專用).xlsx')
CSV_PRICING = BASE / 'NCH定價-2026-06-21.csv'
OUTPUT = BASE / 'nch-seat-map.csv'

SHEET = '國家音樂廳空白圖'


def load_pricing(csv_path: Path) -> set:
    """Return set of (section, row_str, seat_str) from pricing CSV."""
    known = set()
    with open(csv_path, encoding='utf-8-sig') as f:
        lines = [l for l in f if not l.startswith('#')]
    reader = csv.DictReader(lines)
    for row in reader:
        known.add((row['section'].strip(), row['row'].strip(), row['seat'].strip()))
    return known


def int_val(v):
    """Return positive integer value of v, or None if not a positive integer."""
    if isinstance(v, int) and v > 0:
        return v
    if isinstance(v, float) and v == int(v) and v > 0:
        return int(v)
    return None


def main():
    print(f'Reading: {EXCEL_SRC}')
    wb = openpyxl.load_workbook(EXCEL_SRC, data_only=True)
    ws = wb[SHEET]

    print(f'Loading pricing CSV: {CSV_PRICING}')
    known_seats = load_pricing(CSV_PRICING)
    print(f'Known seats: {len(known_seats)}')

    records = []  # (excel_row, col, section, row_str, seat_str)
    unmatched = []

    def record(excel_row, col, section, row_str, seat_str):
        if (section, row_str, seat_str) in known_seats:
            records.append((excel_row, col, section, row_str, seat_str))
        else:
            unmatched.append(f'  no match: section={section} row={row_str} seat={seat_str} excel_row={excel_row} col={col}')

    # ── Constants mirrored from color_nch.py ──────────────────────────────────
    all_2f_sections = {'2樓', '2樓輪椅席', '2樓輪陪席', '2樓友善席', '2樓友陪席'}

    VA_VB_SEAT_COLS = {31, 32, 37, 38, 39, 40, 41, 42, 46, 47}
    VA_VB_LABEL_COLS = {35, 44}

    EXCLUDE_FROM_2F = {17, 15, 64, 10, 12, 14, 65, 67, 69, 5, 7, 73, 75, 3, 77, 62}

    GALLERY_3F_ROW_RANGES = [
        (range(20, 32), '1E'),
        (range(33, 39), '1D'),
        (range(40, 46), '1C'),
        (range(47, 53), '1B'),
    ]
    GALLERY_3F_LEFT_COLS = {10, 12, 14}
    GALLERY_3F_RIGHT_COLS = {65, 67, 69}
    GALLERY_3F_ALL_COLS = GALLERY_3F_LEFT_COLS | GALLERY_3F_RIGHT_COLS

    THREE_F_ROW1_3_EXCEL = {54: '1', 55: '2', 56: '3'}
    EXCLUDE_3F_NUMERIC_1_3 = {21, 58, 5, 7, 73, 75, 10, 12, 67, 69}

    EXCLUDE_3F_NUMERIC_4_11 = {
        29, 50,
        5, 7, 73, 75, 3, 77,
        8, 72,
    }

    GALLERY_4F_LABEL_LEFT = {45: '2E', 52: '2D', 59: '2C', 67: '2B'}
    GALLERY_4F_LABEL_RIGHT = {46: '2E', 52: '2D', 59: '2C', 67: '2B'}
    GALLERY_4F_SEAT_COLS_INNER = {5, 7, 73, 75}
    GALLERY_4F_SEAT_COLS_OUTER = {3, 77}
    ALL_4F_GALLERY_COLS = GALLERY_4F_SEAT_COLS_INNER | GALLERY_4F_SEAT_COLS_OUTER

    EXCLUDE_4F_NUMERIC_1_4 = {
        19, 61,
        5, 7, 73, 75, 3, 77,
        8, 72,
    }
    EXCLUDE_4F_NUMERIC_5_12 = {
        28, 51,
        5, 7, 73, 75, 3, 77,
        8, 72,
    }

    # ── 1. 2F Q-column map ────────────────────────────────────────────────────
    q_map = {}
    for r in range(11, 43):
        v = ws.cell(row=r, column=17).value
        if isinstance(v, int):
            q_map[r] = str(v)

    # ── 2. 2F main floor (excel rows 11-42) ──────────────────────────────────
    for excel_row, row_str in q_map.items():
        for c in range(1, 78):
            if c in EXCLUDE_FROM_2F:
                continue
            v = int_val(ws.cell(row=excel_row, column=c).value)
            if v is None:
                continue
            seat_str = str(v)
            # Try all 2F sections; prefer special types over regular
            best_section = None
            for section in all_2f_sections:
                if (section, row_str, seat_str) in known_seats:
                    if best_section is None:
                        best_section = section
                    elif section != '2樓':
                        best_section = section  # prefer special sections
            if best_section:
                record(excel_row, c, best_section, row_str, seat_str)

    # ── 3. 3F gallery rows (1B, 1C, 1D, 1E) ─────────────────────────────────
    def find_3f_gallery_label(excel_row):
        for row_range, label in GALLERY_3F_ROW_RANGES:
            if excel_row in row_range:
                return label
        return None

    for r in range(15, 53):
        for c in GALLERY_3F_ALL_COLS:
            v = int_val(ws.cell(row=r, column=c).value)
            if v is None:
                continue
            row_label = find_3f_gallery_label(r)
            if row_label is None:
                unmatched.append(f'  3F gallery no label: excel_row={r} col={c} val={v}')
                continue
            record(r, c, '3樓', row_label, str(v))

    # ── 4. 3F numeric rows 1-3 (excel rows 54-56) ────────────────────────────
    for excel_row, row_str in THREE_F_ROW1_3_EXCEL.items():
        for c in range(1, 78):
            if c in EXCLUDE_3F_NUMERIC_1_3:
                continue
            v = int_val(ws.cell(row=excel_row, column=c).value)
            if v is None:
                continue
            record(excel_row, c, '3樓', row_str, str(v))

    # ── 5. 3F numeric rows 4-11 (excel rows 58-65) ───────────────────────────
    THREE_F_ROW_4_11 = {}
    for r in range(58, 66):
        v = ws.cell(row=r, column=29).value
        if isinstance(v, int):
            THREE_F_ROW_4_11[r] = str(v)

    for excel_row, row_str in THREE_F_ROW_4_11.items():
        is_vavb_row = excel_row in (64, 65)
        for c in range(1, 78):
            if c in EXCLUDE_3F_NUMERIC_4_11:
                continue
            if is_vavb_row and c in VA_VB_LABEL_COLS:
                continue
            if is_vavb_row and c in VA_VB_SEAT_COLS:
                continue
            v = int_val(ws.cell(row=excel_row, column=c).value)
            if v is None:
                continue
            record(excel_row, c, '3樓', row_str, str(v))

    # ── 6. VA/VB rows (64-65) ────────────────────────────────────────────────
    VA_VB_ROWS = {64: 'VA', 65: 'VB'}
    for excel_row, row_label in VA_VB_ROWS.items():
        for c in VA_VB_SEAT_COLS:
            v = int_val(ws.cell(row=excel_row, column=c).value)
            if v is None:
                continue
            record(excel_row, c, '3樓', row_label, str(v))

    # ── 7. 4F gallery (2B, 2C, 2D, 2E) ──────────────────────────────────────
    def find_nearest_4f_label(excel_row, side):
        label_map = GALLERY_4F_LABEL_LEFT if side == 'left' else GALLERY_4F_LABEL_RIGHT
        best_dist = 999
        best_label = None
        for label_row, label in label_map.items():
            dist = abs(excel_row - label_row)
            if dist < best_dist:
                best_dist = dist
                best_label = label
        return best_label if best_dist <= 12 else None

    for r in range(40, 75):
        for c in ALL_4F_GALLERY_COLS:
            v = int_val(ws.cell(row=r, column=c).value)
            if v is None:
                continue
            side = 'left' if c in {3, 5, 7} else 'right'
            row_label = find_nearest_4f_label(r, side)
            if row_label is None:
                unmatched.append(f'  4F gallery no label: excel_row={r} col={c} val={v}')
                continue
            record(r, c, '4樓', row_label, str(v))

    # ── 8. 4F numeric rows 1-4 (excel rows 71-74) ────────────────────────────
    FOUR_F_ROW_1_4 = {}
    for r in range(71, 75):
        v = ws.cell(row=r, column=19).value
        if isinstance(v, int):
            FOUR_F_ROW_1_4[r] = str(v)

    for excel_row, row_str in FOUR_F_ROW_1_4.items():
        for c in range(1, 78):
            if c in EXCLUDE_4F_NUMERIC_1_4:
                continue
            v = int_val(ws.cell(row=excel_row, column=c).value)
            if v is None:
                continue
            record(excel_row, c, '4樓', row_str, str(v))

    # ── 9. 4F numeric rows 5-12 (excel rows 76-83) ───────────────────────────
    FOUR_F_ROW_5_12 = {}
    for r in range(76, 84):
        v = ws.cell(row=r, column=28).value
        if isinstance(v, int):
            FOUR_F_ROW_5_12[r] = str(v)

    for excel_row, row_str in FOUR_F_ROW_5_12.items():
        for c in range(1, 78):
            if c in EXCLUDE_4F_NUMERIC_5_12:
                continue
            v = int_val(ws.cell(row=excel_row, column=c).value)
            if v is None:
                continue
            record(excel_row, c, '4樓', row_str, str(v))

    # ── Write output ──────────────────────────────────────────────────────────
    print(f'\nMatched: {len(records)} cells')
    if unmatched:
        print(f'Unmatched/no-label ({len(unmatched)}):')
        for u in unmatched[:10]:
            print(u)

    with open(OUTPUT, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['excel_row', 'col', 'section', 'row', 'seat'])
        for rec in sorted(records):
            writer.writerow(rec)

    print(f'\nWrote {len(records)} rows to: {OUTPUT}')
    return len(records)


if __name__ == '__main__':
    count = main()
    sys.exit(0)
