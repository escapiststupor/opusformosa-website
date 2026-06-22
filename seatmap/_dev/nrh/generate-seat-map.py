#!/usr/bin/env python3
"""
Generate NRH seat-map CSV from blank Excel.

Reads the blank NRH seating chart and outputs a CSV mapping every
Excel cell position that contains a seat number to its (section, row, seat)
identity as found in the pricing CSV.

Output columns: excel_row,col,section,row,seat
"""

import csv
import sys
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
EXCEL_SRC = Path('/Users/pyen/OpusFormosa/ticketing/2026/0910NRH/國家兩廳院演奏廳座位圖1150303(多元友善席專用).xlsx')
CSV_PRICING = BASE / 'NRH定價-2026-06-21.csv'
OUTPUT = BASE / 'nrh-seat-map.csv'

# ── Layout constants ───────────────────────────────────────────────────────────
SHEET = '空白座位圖'
Q_COL = 17  # Column Q: data-row number for every seat row

# All sections that may appear in the NRH pricing CSV
NRH_SECTIONS = [
    '地下樓BF',
    '地下樓BF輪椅席',
    '地下樓BF輪陪席',
    '地下樓BF友善席',
    '地下樓BF友陪席',
]


def load_pricing(csv_path: Path) -> set:
    """
    Return a set of (section, row_str, seat_str) for all seats in the pricing CSV.
    Used to validate/match Excel cells.
    """
    known = set()
    with open(csv_path, encoding='utf-8-sig') as f:
        lines = [l for l in f if not l.startswith('#')]
    reader = csv.DictReader(lines)
    for row in reader:
        known.add((row['section'].strip(), row['row'].strip(), row['seat'].strip()))
    return known


def main():
    print(f'Reading: {EXCEL_SRC}')
    wb = openpyxl.load_workbook(EXCEL_SRC, data_only=True)
    ws = wb[SHEET]

    print(f'Loading pricing CSV: {CSV_PRICING}')
    known_seats = load_pricing(CSV_PRICING)
    print(f'Known seats: {len(known_seats)}')

    # Build row map: excel_row → data_row_str from column Q
    row_map = {}
    for excel_row in range(1, ws.max_row + 1):
        cell = ws.cell(row=excel_row, column=Q_COL)
        if isinstance(cell.value, (int, float)) and cell.value > 0:
            row_map[excel_row] = str(int(cell.value))
    print(f'Row map ({len(row_map)} rows): {row_map}')

    # Scan every seat cell; match against all sections
    records = []
    unmatched = []

    for excel_row, row_str in row_map.items():
        for col in range(1, ws.max_column + 1):
            if col == Q_COL:
                continue
            val = ws.cell(row=excel_row, column=col).value
            if not isinstance(val, (int, float)):
                continue
            seat_str = str(int(val))

            # Find which section this (row, seat) belongs to
            matched_section = None
            for section in NRH_SECTIONS:
                if (section, row_str, seat_str) in known_seats:
                    matched_section = section
                    break

            if matched_section:
                records.append((excel_row, col, matched_section, row_str, seat_str))
            else:
                unmatched.append((excel_row, col, row_str, seat_str))

    print(f'\nMatched: {len(records)} cells')
    if unmatched:
        print(f'Unmatched Excel cells (in pricing CSV but not found): {len(unmatched)}')
        for u in unmatched[:10]:
            print(f'  excel_row={u[0]} col={u[1]} row={u[2]} seat={u[3]}')

    # Write output CSV
    with open(OUTPUT, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['excel_row', 'col', 'section', 'row', 'seat'])
        for rec in sorted(records):
            writer.writerow(rec)

    print(f'\nWrote {len(records)} rows to: {OUTPUT}')
    return len(records)


if __name__ == '__main__':
    count = main()
    sys.exit(0)
