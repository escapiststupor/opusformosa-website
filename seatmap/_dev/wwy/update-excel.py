#!/usr/bin/env python3
"""
根據定價 CSV 在 Excel 座位圖裡新增「定價紀錄」工作表
用法: python3 update-excel.py <csv檔案>
範例: python3 update-excel.py WWY定價-2026-04-03.csv

需要安裝: pip install openpyxl

Excel 效果：
  原始工作表不變。
  新增（或更新）一張「定價紀錄」工作表：
    A: 區域  B: 排  C: 座位  D: 類型  E: 票價  F: 備註
  新增（或更新）一張「票價摘要」工作表。
"""

import csv, sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print('請先安裝 openpyxl：  pip install openpyxl')
    sys.exit(1)

XLSX_PATH = Path('衛武營票圖_音樂廳.xlsx')

TYPE_LABEL_ZH = {
    'regular':    '一般席',
    'staff':      '工作席',
    'console':    '控台區',
    'chorus':     '合唱席',
    'reserved':   '保留席',
    'wheelchair': '輪椅席',
    'companion':  '輪椅陪同席',
    'obstructed': '視線不良席',
}

NON_SALE = {'staff', 'chorus', 'reserved'}

# 票價對應儲存格背景色（Excel hex，無 #）
PRICE_FILLS = [
    'FF6B6B','FFD43B','4DA6FF','51CF66','F783AC',
    'FF922B','74C0FC','A9E34B','CC5DE8','20C997',
]


def main(csv_path):
    # ── 讀 CSV ─────────────────────────────────────────────────────
    rows_data = []
    prices = set()
    with open(csv_path, encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            price = int(row['price']) if row.get('price') else None
            rows_data.append({
                'section': row['section'],
                'row':     row['row'],
                'seat':    row['seat'],
                'type':    row.get('type', 'regular'),
                'price':   price,
            })
            if price:
                prices.add(price)

    # 票價對應顏色
    price_fill = {
        p: PRICE_FILLS[i % len(PRICE_FILLS)]
        for i, p in enumerate(sorted(prices, reverse=True))
    }

    # ── 開啟 Excel ──────────────────────────────────────────────────
    wb = openpyxl.load_workbook(XLSX_PATH)

    for sheet in ('定價紀錄', '票價摘要'):
        if sheet in wb.sheetnames:
            del wb[sheet]

    ws = wb.create_sheet('定價紀錄')

    # ── 標題列 ──────────────────────────────────────────────────────
    headers = ['區域', '排', '座位', '類型', '票價（元）', '備註']
    header_fill  = PatternFill('solid', fgColor='1F3864')
    header_font  = Font(bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center')
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = border

    # ── 資料列 ──────────────────────────────────────────────────────
    for r, d in enumerate(rows_data, 2):
        type_zh = TYPE_LABEL_ZH.get(d['type'], d['type'])
        note = '不販售' if d['type'] in NON_SALE else ''

        vals = [d['section'], d['row'], d['seat'], type_zh, d['price'], note]
        bg = None
        if d['price'] and d['price'] in price_fill:
            bg = PatternFill('solid', fgColor=price_fill[d['price']])
        elif d['type'] != 'regular':
            bg = PatternFill('solid', fgColor='F2F2F2')

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border
            if bg and col == 5:
                cell.fill = bg

    # ── 欄寬 ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10

    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = f'A1:F{len(rows_data)+1}'

    # ── 票價摘要工作表 ──────────────────────────────────────────────
    ws2 = wb.create_sheet('票價摘要')
    ws2.cell(1, 1, '票價（元）').font = Font(bold=True)
    ws2.cell(1, 2, '席數').font       = Font(bold=True)
    ws2.cell(1, 3, '小計（元）').font  = Font(bold=True)

    price_counts = {}
    for d in rows_data:
        if d['price']:
            price_counts[d['price']] = price_counts.get(d['price'], 0) + 1

    total_rev = 0
    for r2, (p, cnt) in enumerate(sorted(price_counts.items(), reverse=True), 2):
        ws2.cell(r2, 1, p)
        ws2.cell(r2, 2, cnt)
        ws2.cell(r2, 3, p * cnt)
        total_rev += p * cnt

    last = len(price_counts) + 2
    ws2.cell(last, 1, '票房上限').font = Font(bold=True)
    ws2.cell(last, 3, total_rev).font  = Font(bold=True)
    ws2.cell(last+1, 1, f'更新日期：{date.today()}')

    # ── 儲存 ────────────────────────────────────────────────────────
    out_path = XLSX_PATH.with_name(
        XLSX_PATH.stem + f'_定價_{date.today()}.xlsx'
    )
    wb.save(out_path)
    print(f'Done. Wrote {out_path}')
    print(f'  已定價：{sum(price_counts.values())} 席')
    print(f'  票房上限：NT${total_rev:,}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python3 update-excel.py <csv檔案>')
        sys.exit(1)
    main(sys.argv[1])
