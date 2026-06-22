#!/usr/bin/env python3
"""
Generic seat-coloring engine.

Usage:
    python3 apply-excel.py <seat-map.csv> <pricing.csv> <blank.xlsx> <output.xlsx>

The seat-map CSV maps Excel cell positions to seat identities:
    excel_row,col,section,row,seat

The pricing CSV maps seat identities to type and price:
    section,row,seat,type,price

Color logic:
  - If the seat type has a type-specific color → use it
  - Else if a price color exists → use it
  - Else → gray (FFCCCCCC)

Does NOT modify legend cells — kept as a pure coloring engine.
"""

import csv
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

# ── Color scheme ───────────────────────────────────────────────────────────────
PRICE_COLORS = {
    5000: 'FFFF6B6B',
    3600: 'FFFFD43B',
    3300: 'FFFF6B6B',  # WWY console
    2800: 'FF4DA6FF',
    2500: 'FF51CF66',
    2200: 'FFFFD43B',  # NRH tier
    2100: 'FFF783AC',
    1800: 'FFFF922B',
    1500: 'FF74C0FC',
    1200: 'FFA9E34B',
    1000: 'FFF783AC',  # NRH tier
    790:  'FFCC5DE8',
    640:  'FF20C997',
    300:  'FFADB5BD',
}

TYPE_COLORS = {
    'staff':                    'FFFFC000',
    'wheelchair':               'FF00B0F0',
    'companion':                'FF0070C0',
    'accessible':               'FF76923C',
    'accessible_companion':     'FFC2D69B',
    'obstructed':               'FF7F7F7F',
    'organ_obstructed':         'FFFBD4B4',
    'uncomfortable_obstructed': 'FF953734',
    'vip_box':                  'FFFF99CC',
    'chorus':                   'FF595959',
    'reserved':                 'FF8E44AD',
    'console':                  'FFE67E22',
    'photography':              'FFFFC000',
}

FALLBACK_COLOR = 'FFCCCCCC'


def make_fill(argb: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=argb)


def pick_color(seat_type: str, price: int | None) -> str:
    if seat_type in TYPE_COLORS:
        return TYPE_COLORS[seat_type]
    if price and price in PRICE_COLORS:
        return PRICE_COLORS[price]
    return FALLBACK_COLOR


def load_seat_map(path: Path) -> dict:
    """
    Returns {(excel_row, col): (section, row_str, seat_str)}
    """
    mapping = {}
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (int(row['excel_row']), int(row['col']))
            val = (row['section'].strip(), row['row'].strip(), row['seat'].strip())
            mapping[key] = val
    return mapping


def load_pricing(path: Path) -> dict:
    """
    Returns {(section, row_str, seat_str): (seat_type, price_or_None)}
    """
    pricing = {}
    with open(path, encoding='utf-8-sig') as f:
        lines = [l for l in f if not l.startswith('#')]
    reader = csv.DictReader(lines)
    for row in reader:
        key = (row['section'].strip(), row['row'].strip(), row['seat'].strip())
        price_raw = row['price'].strip()
        price = int(price_raw) if price_raw else None
        pricing[key] = (row['type'].strip(), price)
    return pricing


def main():
    if len(sys.argv) != 5:
        print('Usage: apply-excel.py <seat-map.csv> <pricing.csv> <blank.xlsx> <output.xlsx>')
        sys.exit(1)

    seat_map_path = Path(sys.argv[1])
    pricing_path  = Path(sys.argv[2])
    blank_path    = Path(sys.argv[3])
    output_path   = Path(sys.argv[4])

    print(f'Seat map:  {seat_map_path}')
    print(f'Pricing:   {pricing_path}')
    print(f'Blank:     {blank_path}')
    print(f'Output:    {output_path}')

    seat_map = load_seat_map(seat_map_path)
    pricing  = load_pricing(pricing_path)

    print(f'\nSeat map entries: {len(seat_map)}')
    print(f'Pricing entries:  {len(pricing)}')

    # Copy blank → output (never modify original)
    shutil.copy2(blank_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    # Use first sheet if only one, otherwise require caller to know sheet name
    # For multi-sheet workbooks the seat-map already encodes the right cells
    ws = wb.active

    colored_count = 0
    type_stats = {}
    price_stats = {}

    for (excel_row, col), (section, row_str, seat_str) in seat_map.items():
        key = (section, row_str, seat_str)
        if key not in pricing:
            continue
        seat_type, price = pricing[key]
        color = pick_color(seat_type, price)
        ws.cell(row=excel_row, column=col).fill = make_fill(color)
        colored_count += 1
        if seat_type in TYPE_COLORS:
            type_stats[seat_type] = type_stats.get(seat_type, 0) + 1
        else:
            label = f'price:{price}'
            price_stats[label] = price_stats.get(label, 0) + 1

    wb.save(output_path)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f'\nTotal cells colored: {colored_count}')
    print('\nBreakdown by type:')
    for t, c in sorted(type_stats.items()):
        print(f'  {t}: {c}')
    print('\nBreakdown by price:')
    for p, c in sorted(price_stats.items(), key=lambda x: -x[1]):
        print(f'  {p}: {c}')
    print(f'\nSaved: {output_path}')


if __name__ == '__main__':
    main()
